from openpyxl import load_workbook


def parse_excel(file_stream):
    """Parse an .xlsx file. Row 1 = headers. Returns (headers, rows) where rows is a list of dicts."""
    try:
        wb = load_workbook(file_stream, read_only=True, data_only=True)
    except Exception:
        raise ValueError("Cannot read file. Please ensure it is a valid .xlsx format")

    ws = wb.active
    if ws is None:
        raise ValueError("No worksheets found in the Excel file")

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel file is empty")

    headers = []
    for h in header_row:
        val = str(h).strip() if h is not None else ''
        if val:
            headers.append(val)

    if not headers:
        raise ValueError("No column headers found (row 1 is empty)")

    # Data rows
    data_rows = []
    for row in rows_iter:
        record = {}
        all_empty = True
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            cell_val = str(val).strip() if val is not None else ''
            record[headers[i]] = cell_val
            if cell_val:
                all_empty = False
        if not all_empty:
            data_rows.append(record)

    wb.close()
    return headers, data_rows
